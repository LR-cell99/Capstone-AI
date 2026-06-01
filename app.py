import streamlit as st
import openai, os, re, io, hashlib, hashlib
from datetime import date, datetime
from pathlib import Path
import pandas as pd
from dotenv import load_dotenv
from fpdf import FPDF
import pdfplumber, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from docx import Document
from supabase import create_client, Client as SupabaseClient

load_dotenv()
OPENAI_API_KEY    = os.getenv("OPENAI_API_KEY")
SUPABASE_URL      = os.getenv("SUPABASE_URL", "")
SUPABASE_ANON_KEY = os.getenv("SUPABASE_ANON_KEY", "")
BASE_RESUME_DIR   = Path("base_resume")
SUPPORTED_EXTS    = (".pdf", ".docx", ".txt")
STATUS_OPTIONS    = ["Not Applied", "Pending", "Applied", "Interview", "Offer", "Rejected"]

# ── Supabase ───────────────────────────────────────────────────────────────────
@st.cache_resource
def get_supabase() -> SupabaseClient | None:
    if not SUPABASE_URL or not SUPABASE_ANON_KEY:
        return None
    try:
        return create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
    except Exception as e:
        st.error(f"Supabase connection error: {e}")
        return None

def sb_load() -> pd.DataFrame:
    sb = get_supabase()
    empty = pd.DataFrame(columns=["Company", "Role", "Date Applied", "Status", "id"])
    if not sb:
        return empty
    try:
        rows = sb.table("applications").select("*").order("created_at").execute().data or []
        if not rows:
            return empty
        df = pd.DataFrame(rows).rename(columns={
            "company": "Company", "role": "Role",
            "date_applied": "Date Applied", "status": "Status"})
        return df[[c for c in ["Company", "Role", "Date Applied", "Status", "id"] if c in df.columns]]
    except Exception:
        return empty

def sb_insert(company, role, date_applied, status) -> tuple[str | None, str | None]:
    sb = get_supabase()
    if not sb:
        return None, "Supabase not initialised — check .env"
    try:
        d = date_applied.strftime("%Y-%m-%d") if hasattr(date_applied, "strftime") else str(date_applied)[:10]
        res = sb.table("applications").insert(
            {"company": str(company), "role": str(role), "date_applied": d, "status": str(status)}
        ).execute()
        return (res.data[0]["id"], None) if res.data else (None, "Insert returned no data")
    except Exception as e:
        return None, str(e)

def sb_update(row_id, company, role, date_applied, status):
    sb = get_supabase()
    if not sb or not row_id:
        return
    try:
        sb.table("applications").update({
            "company": company, "role": role,
            "date_applied": str(date_applied), "status": status
        }).eq("id", row_id).execute()
    except Exception:
        pass

# ── Text extraction ────────────────────────────────────────────────────────────
def extract_pdf(file) -> str:
    with pdfplumber.open(file) as pdf:
        return "\n".join(p.extract_text() for p in pdf.pages if p.extract_text()).strip()

def extract_docx(file) -> str:
    return "\n".join(p.text for p in Document(file).paragraphs if p.text.strip())

def extract_file(f) -> str:
    b = io.BytesIO(f.read())
    n = f.name.lower()
    if n.endswith(".pdf"):   return extract_pdf(b)
    if n.endswith(".docx"):  return extract_docx(b)
    b.seek(0); return b.read().decode("utf-8", errors="ignore").strip()

def load_base_resume() -> tuple[str, str]:
    if not BASE_RESUME_DIR.exists():
        return "", ""
    for ext in SUPPORTED_EXTS:
        matches = list(BASE_RESUME_DIR.glob(f"*{ext}"))
        if matches:
            fp = matches[0]
            b = io.BytesIO(fp.read_bytes())
            text = extract_pdf(b) if ext == ".pdf" else extract_docx(b) if ext == ".docx" else fp.read_text(encoding="utf-8", errors="ignore").strip()
            return text, fp.name
    return "", ""

# ── PDF export ─────────────────────────────────────────────────────────────────
_PDF_CHARS = {"\u2013":"-","\u2014":"-","\u2012":"-","\u2015":"-","\u2018":"'",
              "\u2019":"'","\u201c":'"',"\u201d":'"',"\u2022":"-","\u2026":"...","\u00a0":" "}

def sanitise(text: str) -> str:
    for c, r in _PDF_CHARS.items():
        text = text.replace(c, r)
    return text.encode("latin-1", errors="ignore").decode("latin-1")

def generate_pdf(text: str) -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_margins(20, 20, 20)
    pdf.set_auto_page_break(True, 20)
    for line in text.split("\n"):
        s = sanitise(line.strip())
        if not s:
            pdf.ln(3); continue
        if s.isupper() and len(s) > 2:
            pdf.set_font("Helvetica", "B", 13); pdf.ln(3); pdf.cell(0, 8, s, ln=True)
            pdf.set_draw_color(45, 90, 61); pdf.set_line_width(0.5)
            pdf.line(20, pdf.get_y(), 190, pdf.get_y()); pdf.ln(2)
        elif s.startswith(("-", "*")):
            pdf.set_font("Helvetica", "", 10); pdf.set_x(25)
            pdf.cell(5, 6, "-", ln=False); pdf.multi_cell(0, 6, s.lstrip("-* ").strip())
        elif len(s) < 60 and not s.endswith("."):
            pdf.set_font("Helvetica", "B", 10); pdf.cell(0, 6, s, ln=True)
        else:
            pdf.set_font("Helvetica", "", 10); pdf.multi_cell(0, 6, s)
    return bytes(pdf.output())

# ── Excel export ───────────────────────────────────────────────────────────────
def export_excel(df: pd.DataFrame) -> bytes:
    df = df.drop(columns=["#"], errors="ignore").copy()
    out = io.BytesIO()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Application Tracker"
    hfill = PatternFill("solid", fgColor="2D5A3D")
    hfont = Font(bold=True, color="FFFFFF", size=11)
    halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col); c.font = hfont; c.fill = hfill
        c.alignment = halign; c.border = border
    ws.row_dimensions[1].height = 20
    afill = PatternFill("solid", fgColor="E8F2EB")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, str(val) if val is not None else "")
            c.alignment = Alignment(vertical="center", wrap_text=True); c.border = border
            if ri % 2 == 0: c.fill = afill
        ws.row_dimensions[ri].height = 18
    for ci, col in enumerate(df.columns, 1):
        mx = max([len(str(col))] + [max(len(l) for l in str(v).split("\n")) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(ci)].width = min(max(mx + 4, 12), 60)
    ws.freeze_panes = "A2"; wb.save(out); return out.getvalue()

# ── JD helpers ─────────────────────────────────────────────────────────────────
_NOISE = [r"(?i)by continuing to use our platform", r"(?i)open app", r"(?i)privacy policy",
          r"(?i)cookie.*consent", r"(?i)your application will include", r"(?i)employer questions",
          r"(?i)registration no\.", r"(?i)\bEA No\b", r"(?i)company information",
          r"(?i)view all jobs", r"(?i)(full|part) time.*per month",
          r"(?i)\$[\d,]+\s*[-\u2013]\s*\$[\d,]+\s*per\s*(month|year|hour)",
          r"(?i)posted \d+\w* ago", r"(?i)(medium|high|low) application volume",
          r"(?i)(benefits.*allowance|variable bonus|apply now|quick apply|save job|report job|share this job|similar jobs|you might also like|javascript is disabled)",
          r"(?i)(central|north|south|east|west) region"]

def clean_jd(raw: str) -> str:
    lines = [l.strip() for l in raw.split("\n")
             if not any(re.search(p, l.strip()) for p in _NOISE)]
    return re.sub(r"\n{3,}", "\n\n", "\n".join(lines)).strip()

def clean_field_value(val: str) -> str:
    """Strip parenthetical expansions and extra whitespace from extracted field values.
    e.g. 'SSMC (Systems on Silicon Manufacturing Company Pte. Ltd.)' -> 'SSMC'
    """
    val = re.sub(r"\s*\(.*?\)", "", val).strip()  # Remove anything in parentheses
    val = re.sub(r"\s{2,}", " ", val)               # Collapse multiple spaces
    return val.strip()

def parse_jd_field(text: str, labels: list) -> str:
    lines = text.split("\n")
    for i, line in enumerate(lines):
        line = line.strip()
        for pat in labels:
            if re.match(pat + r"\s*[:\-]\s*.+", line, re.I):
                v = re.sub(pat + r"\s*[:\-]\s*", "", line, flags=re.I).strip()
                v = clean_field_value(v)
                if v and v.lower() not in ("na", "not applicable", "none", ""):
                    return v
            if re.match(pat + r"\s*[:\-]?\s*$", line, re.I):
                for nxt in lines[i+1:i+3]:
                    nxt = clean_field_value(nxt.strip())
                    if nxt and not re.match(r"(?i)^(job title|role|position|key responsibilities|required|preferred|company name|company|about)", nxt):
                        if nxt.lower() not in ("na", "not applicable", "none", ""):
                            return nxt
                        break
    return "NA"

# ── Prompts ────────────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """You are a senior professional resume writer for the Singapore job market. Edit the candidate's resume to better align it with the job description — not a full rewrite.

GOAL: The enhanced resume must feel like the same person wrote it, with sharper language, better JD alignment, and stronger framing. Preserve the candidate's authentic voice and actual experience. Only change what needs changing — but everything that can be aligned to the JD, should be.

STEP 1 — ANALYSE BEFORE EDITING:
- Read every section of the resume fully.
- Extract from the JD: role title, key responsibilities, required skills, preferred skills, industry context, and any repeated terminology.
- For every section of the resume, ask: does the language, framing, and emphasis reflect what this JD is looking for?
- Note what is already strong and JD-aligned — leave those parts alone.
- Identify everything that can be reworded, reframed, or reordered to better reflect the JD — without adding anything that isn't there.

STEP 2 — TARGETED EDITING BY SECTION:

EXECUTIVE SUMMARY / PROFESSIONAL PROFILE:
- Use the original as your foundation — do not discard the candidate's intent.
- Rewrite to open with the JD's exact job title or a close variant.
- Naturally embed 3-5 JD keywords and the industry context (e.g. semiconductor, manufacturing, automation).
- Ensure the tone and focus reflects what the JD prioritises — if the JD emphasises cross-functional collaboration, the summary should mention it; if it emphasises technical depth, lead with that.
- Keep the same general length. The candidate should recognise their own summary.

WORK EXPERIENCE BULLETS:
- Bullets already strong and JD-relevant — keep with light polish only.
- Vague, passive, or JD-misaligned bullets — rewrite using: [Strong verb] + [what you did] + [how/with what] + [result or impact].
- Use the JD's exact terminology where applicable (e.g. "root cause analysis" not "finding problems", "preventive maintenance" not "machine upkeep").
- Add a reasonable result where none exists and context supports it — do not fabricate numbers not supported by the original.
- Strong verbs: Executed, Implemented, Collaborated, Optimised, Reduced, Troubleshot, Calibrated, Validated, Coordinated, Monitored, Analysed, Resolved, Streamlined, Facilitated, Engineered, Delivered, Spearheaded.

PROJECTS:
- If the resume has a projects section, reframe each project description to emphasise aspects most relevant to the JD.
- Highlight technologies, methodologies, or outcomes that align with what the JD requires.
- If a project used a skill the JD values, make that skill prominent in the description.
- Do not change the project name, tech stack, or outcomes — only the framing and emphasis.

ACHIEVEMENTS / AWARDS / CERTIFICATIONS:
- Reorder to place JD-relevant achievements first.
- Reword achievement descriptions to use JD language where applicable.
- Keep all facts intact — only rephrase, never fabricate.

SKILLS:
- Reorder so JD-critical skills appear first.
- Where a skill is named differently from the JD equivalent, use the JD's exact terminology.
- Only add a skill if the candidate's work experience or projects clearly demonstrate it.

EDUCATION:
- Keep factually intact. No changes unless a certification or qualification directly matches a JD requirement — in that case, make sure it is prominently visible.

HALLUCINATION RULES — non-negotiable:
- NEVER invent a job, company, project, degree, certification, or skill not in the original resume.
- NEVER fabricate metrics, outcomes, or achievements.
- Reframing, rewording, reordering, and re-emphasising existing content is encouraged and is NOT hallucination.
- You may make implied results explicit if the original context supports them.

OUTPUT RULES: Output ONLY the resume. No preamble, commentary, or explanation. No markdown. First character = candidate name. Do not add any closing line or marker after the last line of resume content."""

JD_EXTRACT_PROMPT = """You are a strict job description parser for Singapore job portals. Extract ONLY what a hiring manager actually wrote.

EXTRACT:
1. Company name — actual employer only. NOT recruiter agencies or applicant labels. Leave blank if uncertain.
2. Job title
3. About the company — only if genuine employer description exists
4. Key responsibilities
5. Required qualifications and skills
6. Preferred qualifications
7. Any other role-relevant content

STRIP everything else: salary, allowances, CPF, EA/registration numbers, application instructions, employer questions, cookie/privacy notices, portal UI ("Open app", "Save job", "Quick apply"), applicant tags ("Strong applicant", "Posted X ago"), recruiter boilerplate, unrelated listings, region tags.

OUTPUT: Clean plain text with section labels. Skip missing sections. When in doubt — leave it out."""

_ATS_RULES = """
STEP 0 — VALIDATE JD: If JD lacks real hiring content (Wikipedia, generic pages, all "Not applicable"), set all scores to 0, SUMMARY to "Invalid job description", IMPROVEMENTS to "Upload a real job posting", and stop.

SCORING (valid JD only):
- There is NO minimum and NO maximum anchor. Scores range freely from 0 to 100.
- A resume that is an excellent match for the JD SHOULD score 80-95+. Do not hold back high scores when genuinely earned.
- 60+ = PASS. Below 60 = FAIL. But a great resume can score 75, 85, or 95 — score what you actually see.
- Score only what is in the resume. Do NOT give benefit of the doubt for weak content. Do NOT hold back for strong content.

KEYWORD MATCH (30 pts):
- 0-10: Missing more than half the critical JD keywords
- 11-18: Most keywords loosely present, some critical ones missing
- 19-24: Strong keyword presence, minor gaps only
- 25-30: Excellent — nearly all JD keywords and SkillsFuture terms present, naturally integrated

RELEVANCE OF EXPERIENCE (25 pts):
- 0-8: Different field or only loosely related
- 9-15: Some relevant experience, significant gaps
- 16-20: Experience clearly maps to role responsibilities
- 21-25: Excellent — experience directly and comprehensively addresses the JD responsibilities

QUALIFICATIONS MATCH (20 pts):
- 0-8: Missing required degree or certification, wrong field
- 9-13: Partially meets requirements
- 14-17: Meets core requirements
- 18-20: Excellent — fully meets or exceeds all stated qualifications and preferred requirements

RESUME CLARITY & STRUCTURE (15 pts):
- 0-5: Vague, passive language, no action verbs, no results
- 6-9: Some strong language but inconsistent
- 10-12: Consistently strong language, good structure
- 13-15: Excellent — strong action verbs throughout, quantified results, tightly targeted summary

ATS FORMATTING COMPLIANCE (10 pts):
- 0-4: Tables, columns, or special characters present
- 5-7: Minor formatting issues
- 8-10: Clean plain text, standard headers, consistent dates

TOTAL_SCORE = exact sum of all five categories.

Respond ONLY in this format:
TOTAL_SCORE: <n>
KEYWORD_MATCH: <n>/30
RELEVANCE_OF_EXPERIENCE: <n>/25
QUALIFICATIONS_MATCH: <n>/20
RESUME_CLARITY: <n>/15
ATS_FORMATTING: <n>/10
SUMMARY: <2-3 sentences — be specific, name actual strengths AND weaknesses, name missing keywords if any>
IMPROVEMENTS: <3 specific actionable bullet points referencing exact JD requirements>"""

BASELINE_ATS_PROMPT = (
    "You are a precise ATS evaluator for Singapore. Score the ORIGINAL BASELINE resume honestly. "
    "This is the benchmark — score it for what it actually is, not what it could be. "
    "A strong baseline resume against a well-matched JD can score 70, 80, or higher. Score freely across the full 0-100 range."
) + _ATS_RULES
BASELINE_ATS_TEMP   = 0.0

ENHANCED_ATS_PROMPT = (
    "You are a precise ATS evaluator for Singapore. Score this AI-ENHANCED resume against the JD. "
    "Score freely across the full 0-100 range — do not anchor to 60 or any other number. "
    "A resume that comprehensively addresses the JD with strong keywords, relevant experience, and clear structure SHOULD score 75-95+. "
    "Score what you see: reward genuine alignment, penalise genuine gaps."
) + _ATS_RULES
ENHANCED_ATS_TEMP   = 0.0

RETRY_ATS_PROMPT = (
    "You are a strict ATS evaluator for Singapore. Quick regression check — score this resume against the JD. "
    "Penalise vague or generic resumes. Reward strong JD alignment. Score freely from 0-100."
) + _ATS_RULES
RETRY_ATS_TEMP      = 0.0

SKILLS_GAP_PROMPT = """You are a career coach and skills analyst for the Singapore job market.

You will receive a resume and a job description. Analyse the skills match and return a structured gap analysis.

OUTPUT ONLY in this exact format — no extra text:

HAVE:
- <skill or competency from JD that the resume clearly demonstrates>
(repeat for each)

MISSING:
- <skill or competency required by the JD that is absent from the resume>
(repeat for each)

IRRELEVANT:
- <skill on the resume that is not mentioned or useful for this role>
(repeat for each)

Rules:
- Be specific — name actual skills, tools, methodologies, not vague categories
- HAVE: only include if clearly evidenced in the resume, not just implied
- MISSING: only include if explicitly required or strongly preferred in the JD
- IRRELEVANT: only include if clearly unrelated to the role
- If a section has nothing to list, write "None identified" under it
- Do not invent skills not mentioned in either document"""

REVISION_PROMPT = """You are a professional resume editor. You will receive:
1. The current enhanced resume
2. The original job description it was tailored for
3. A specific revision request from the candidate

Apply the requested revision carefully. Maintain the same structure and formatting.

HALLUCINATION RULES: NEVER add skills, jobs, degrees, or certifications not in the resume.
OUTPUT RULES: Output ONLY the revised resume. No preamble, no commentary, no explanation. Do not add any closing line, note, or marker after the last line of resume content."""

# ── Session state ──────────────────────────────────────────────────────────────
_SS_DEFAULTS = {
    "enhanced_resume": "", "edited_resume": "", "_last_enhanced": "",
    "jd_text": "", "jd_filename": "", "jd_editable_value": "",
    "jd_widget_version": 0, "ats_result": None, "baseline_ats_result": None,
    "enhance_count": 0, "is_enhancing": False,
    "skills_gap": None, "revision_history": [], "revision_version": 0,
    "resume_versions": [], "ats_score_cache": {}, "active_version_idx": 0,
    "resume_versions": [], "ats_score_cache": {}, "active_version_idx": 0, "revision_version": 0,
}
for k, v in _SS_DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

if "tracker_edit" not in st.session_state:
    st.session_state.tracker_edit = sb_load()

if "resume_text" not in st.session_state or not st.session_state.resume_text:
    _t, _n = load_base_resume()
    st.session_state.resume_text = _t
    st.session_state.base_resume_name = _n
elif "base_resume_name" not in st.session_state:
    st.session_state.base_resume_name = ""

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="AI Resume Enhancer", page_icon="📄", layout="wide")
st.markdown("""<style>
    .block-container{padding-top:2rem;padding-bottom:3rem}
    .stTextArea textarea{font-family:'DM Mono',monospace;font-size:13px;border-radius:10px}
    .stButton>button{background-color:#2d5a3d;color:white;border:none;border-radius:10px;padding:.5rem 1.5rem;font-weight:500}
    .stButton>button:hover{background-color:#214a2e}
    .warn-box{border:1px solid #e6c87a;border-radius:10px;padding:.75rem 1rem;font-size:13px;margin:1rem 0;background:#fdf6e3;color:#7a5a10}
    @media(prefers-color-scheme:dark){.warn-box{background:#2a2200;color:#f0c040;border-color:#7a5a10}}
    .output-box{border:1px solid #e0dbd3;border-radius:12px;padding:1.25rem;font-family:'DM Mono',monospace;font-size:13px;line-height:1.8;white-space:pre-wrap;min-height:200px;background:transparent;color:#bbb5ad}
</style>""", unsafe_allow_html=True)

st.markdown("# 📄 AI Resume Enhancer")
st.caption("Upload a JD — your base resume loads automatically from base_resume/.")
st.divider()

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")
    if OPENAI_API_KEY:
        st.success("✅ OpenAI API key loaded")
    else:
        st.error("❌ OPENAI_API_KEY not found in .env")
    if SUPABASE_URL and SUPABASE_ANON_KEY:
        st.success("✅ Supabase connected")
    else:
        st.warning("⚠️ Supabase not configured — add SUPABASE_URL and SUPABASE_ANON_KEY to .env")
    model = st.selectbox("Model", ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"],
                         help="gpt-4o gives the best results.")
    st.divider()
    output_dir = st.text_input("Save resumes to folder:", value="resume/")
    st.divider()
    st.markdown("**File types:** PDF, DOCX, TXT\n\n**Tip:** Use full JDs from company career pages for best results.")

# ── Upload panels ──────────────────────────────────────────────────────────────
col_resume, col_jd = st.columns(2)

with col_resume:
    st.subheader("📎 Base Resume")
    if st.session_state.resume_text:
        n = st.session_state.base_resume_name or "manually entered"
        st.success(f"✅ Loaded **{n}** ({len(st.session_state.resume_text.split())} words)")
        with st.expander("Preview base resume"):
            st.text(st.session_state.resume_text[:1500] + ("…" if len(st.session_state.resume_text) > 1500 else ""))
        if st.button("🔄 Reload from base_resume/"):
            t, n = load_base_resume()
            if t:
                st.session_state.resume_text = t; st.session_state.base_resume_name = n; st.rerun()
            else:
                st.error("No resume found in base_resume/.")
    else:
        st.warning("⚠️ No resume found in `base_resume/` folder.")
        st.caption("Place your resume (PDF, DOCX, or TXT) in `base_resume/` next to app.py.")
    with st.expander("Override: upload a different resume"):
        rf = st.file_uploader("Upload resume", type=["pdf","docx","txt"], label_visibility="collapsed", key="resume_upload")
        if rf:
            t = extract_file(rf)
            if t:
                st.session_state.resume_text = t
                st.session_state.base_resume_name = rf.name
                # New resume — clear version history and all results
                st.session_state.resume_versions = []
                st.session_state.ats_score_cache = {}
                st.session_state.active_version_idx = 0
                st.session_state.enhanced_resume = ""
                st.session_state.edited_resume = ""
                st.session_state.ats_result = None
                st.session_state.baseline_ats_result = None
                st.session_state.skills_gap = None
                st.session_state.revision_history = []
                st.session_state.revision_version = 0
                st.success(f"✅ Using **{rf.name}** — previous enhancements cleared.")
            else:
                st.error("Could not extract text.")

with col_jd:
    st.subheader("📋 Job Description")
    jf = st.file_uploader("Upload JD", type=["pdf","docx","txt"], label_visibility="collapsed", key="jd_upload")
    if jf and st.session_state.jd_filename != jf.name:
        with st.spinner("Extracting and filtering JD…"):
            raw = extract_file(jf)
        if raw:
            pre = clean_jd(raw)
            try:
                r = openai.OpenAI(api_key=OPENAI_API_KEY).chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role":"system","content":JD_EXTRACT_PROMPT},{"role":"user","content":pre}],
                    temperature=0.0, max_tokens=1000)
                ai_jd = r.choices[0].message.content.strip()
            except Exception:
                ai_jd = pre
            st.session_state.jd_text = ai_jd
            st.session_state.jd_editable_value = ai_jd
            st.session_state.jd_filename = jf.name
            st.session_state.jd_widget_version += 1
            # New JD — clear version history and all results
            st.session_state.resume_versions = []
            st.session_state.ats_score_cache = {}
            st.session_state.active_version_idx = 0
            st.session_state.enhanced_resume = ""
            st.session_state.edited_resume = ""
            st.session_state.ats_result = None
            st.session_state.baseline_ats_result = None
            st.session_state.skills_gap = None
            st.session_state.revision_history = []
            st.session_state.revision_version = 0
            st.success(f"✅ JD processed from **{jf.name}**")
        else:
            st.error("Could not extract text.")

    if st.session_state.jd_editable_value:
        wk = f"jd_editable_{st.session_state.jd_widget_version}"
        if st.session_state.is_enhancing:
            st.caption("🔒 JD locked during enhancement.")
            st.text_area("JD", label_visibility="collapsed", value=st.session_state.jd_editable_value,
                         height=250, key=wk+"_locked", disabled=True)
        else:
            st.caption("✏️ Edit to correct company name, job role, or missing info before enhancing.")
            ejd = st.text_area("JD", label_visibility="collapsed", value=st.session_state.jd_editable_value,
                               height=250, key=wk, help="Edits here flow into enhancement and tracker.")
            st.session_state.jd_text = ejd
            st.session_state.jd_editable_value = ejd
    elif not st.session_state.jd_editable_value:
        with st.expander("Or paste JD manually"):
            mjd = st.text_area("JD text", label_visibility="collapsed",
                               placeholder="Paste the job description here…", height=200, key="manual_jd")
            if mjd.strip():
                st.session_state.jd_text = mjd; st.session_state.jd_editable_value = mjd

# ── Enhance button ─────────────────────────────────────────────────────────────
st.write("")
col_btn, col_clear = st.columns([1, 5])
with col_btn:
    enhance_clicked = st.button("⚡ Enhance Resume", use_container_width=True)
with col_clear:
    if st.button("Clear all"):
        for k, v in _SS_DEFAULTS.items():
            st.session_state[k] = v
        st.session_state.jd_widget_version += 1
        st.rerun()

if enhance_clicked:
    if not OPENAI_API_KEY:
        st.error("OPENAI_API_KEY not found.")
    elif not st.session_state.resume_text.strip():
        st.error("Please upload a resume.")
    elif not st.session_state.jd_text.strip():
        st.error("Please upload a job description.")
    else:
        st.session_state.is_enhancing = True
        client = openai.OpenAI(api_key=OPENAI_API_KEY)

        with st.spinner("Step 1/3 — Scoring baseline resume…"):
            try:
                r = client.chat.completions.create(model="gpt-4o-mini", temperature=BASELINE_ATS_TEMP, max_tokens=600,
                    messages=[{"role":"system","content":BASELINE_ATS_PROMPT},
                              {"role":"user","content":f"## Resume\n\n{st.session_state.resume_text}\n\n---\n\n## JD\n\n{st.session_state.jd_text}"}])
                st.session_state.baseline_ats_result = r.choices[0].message.content.strip()
            except Exception:
                st.session_state.baseline_ats_result = None

        with st.spinner("Step 2/3 — Enhancing your resume…"):
            try:
                def sim(a, b):
                    wa, wb = set(a.lower().split()), set(b.lower().split())
                    return len(wa & wb) / max(len(wa), len(wb)) if wa and wb else 1.0

                uc = f"## Candidate Resume\n\n{st.session_state.resume_text}\n\n---\n\n## Job Description\n\n{st.session_state.jd_text}"
                baseline_total = None
                if st.session_state.baseline_ats_result:
                    m = re.search(r"TOTAL_SCORE:\s*([\d.]+)", st.session_state.baseline_ats_result)
                    if m: baseline_total = float(m.group(1))

                enhanced = None
                for attempt in range(3):
                    temp = [0.7, 0.85, 0.9][attempt]
                    candidate = client.chat.completions.create(model=model, temperature=temp, max_tokens=3500,
                        messages=[{"role":"system","content":SYSTEM_PROMPT},{"role":"user","content":uc}]
                    ).choices[0].message.content.strip()

                    similarity = sim(st.session_state.resume_text, candidate)
                    cand_score = None
                    if baseline_total is not None:
                        try:
                            qr = client.chat.completions.create(model="gpt-4o-mini", temperature=RETRY_ATS_TEMP, max_tokens=600,
                                messages=[{"role":"system","content":RETRY_ATS_PROMPT},
                                          {"role":"user","content":f"## Enhanced Resume\n\n{candidate}\n\n---\n\n## JD\n\n{st.session_state.jd_text}"}])
                            qm = re.search(r"TOTAL_SCORE:\s*([\d.]+)", qr.choices[0].message.content)
                            if qm: cand_score = float(qm.group(1))
                        except Exception:
                            pass

                    regressed = baseline_total and cand_score and cand_score < baseline_total
                    if attempt == 2 or (similarity < 0.80 and not regressed):
                        enhanced = candidate
                        if similarity >= 0.80:
                            st.warning(f"⚠️ Enhancement is very similar to baseline ({similarity:.0%}). JD may have too little content.")
                        elif regressed:
                            st.warning(f"⚠️ Best available result after {attempt+1} attempt(s). Score could not be improved beyond baseline.")
                        break

                st.session_state.enhanced_resume = enhanced
                # Store as version 0 — label includes company and role from JD
                wk0 = f"jd_editable_{st.session_state.jd_widget_version}"
                src0 = st.session_state.get(wk0, st.session_state.jd_text)
                v0_company = parse_jd_field(src0, [r"(?i)company name", r"(?i)company"])
                v0_role    = parse_jd_field(src0, [r"(?i)job title", r"(?i)role", r"(?i)position"])
                v0_label   = f"Original — {v0_company} | {v0_role}" if v0_company != "NA" or v0_role != "NA" else "Original Enhancement"
                st.session_state.resume_versions = [{
                    "label": v0_label,
                    "content": enhanced,
                    "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M"),
                }]
                st.session_state.active_version_idx = 0
                st.session_state.ats_score_cache = {}  # Clear cache for new JD/resume
                Path(output_dir).mkdir(parents=True, exist_ok=True)
                ts = datetime.now().strftime("%Y-%m-%d_%H%M")
                (Path(output_dir) / f"enhanced_{ts}.txt").write_text(enhanced, encoding="utf-8")
                st.success(f"✅ Saved to `{output_dir}enhanced_{ts}.txt`")

                st.session_state.enhance_count += 1
                wk = f"jd_editable_{st.session_state.jd_widget_version}"
                src = st.session_state.get(wk, st.session_state.jd_text)
                company = parse_jd_field(src, [r"(?i)company name", r"(?i)company"])
                role    = parse_jd_field(src, [r"(?i)job title", r"(?i)role", r"(?i)position"])
                st.session_state.tracker_edit = pd.concat([st.session_state.tracker_edit, pd.DataFrame([{
                    "Company": company, "Role": role,
                    "Date Applied": str(date.today()), "Status": "Not Applied", "id": ""
                }])], ignore_index=True)

                # ── Step 3: Skills gap analysis ──
                # (ATS scoring of enhanced resume is now a manual button — see below)
                try:
                    gap_content = f"## Resume\n\n{enhanced}\n\n---\n\n## JD\n\n{st.session_state.jd_text}"
                    gr = client.chat.completions.create(model="gpt-4o-mini", temperature=0.0, max_tokens=800,
                        messages=[{"role":"system","content":SKILLS_GAP_PROMPT},{"role":"user","content":gap_content}])
                    st.session_state.skills_gap = gr.choices[0].message.content.strip()
                except Exception:
                    st.session_state.skills_gap = None

            except openai.AuthenticationError:
                st.error("Invalid API key.")
            except openai.RateLimitError:
                st.error("Rate limit hit — please wait and try again.")
            except Exception as e:
                st.error(f"Error: {e}")
            finally:
                st.session_state.is_enhancing = False

# ── Output ─────────────────────────────────────────────────────────────────────
st.divider()
st.subheader("Enhanced Resume Output")

if st.session_state.enhanced_resume:
    # ── Version switcher ──
    versions = st.session_state.resume_versions
    if len(versions) > 1:
        version_labels = [f"{'🔵' if i == 0 else '✏️'} {v['label']} ({v['timestamp']})"
                          for i, v in enumerate(versions)]
        selected_v = st.selectbox(
            "Select version to view / download:",
            version_labels,
            index=st.session_state.active_version_idx,
            key=f"version_select_{len(versions)}",
        )
        new_idx = version_labels.index(selected_v)
        if new_idx != st.session_state.active_version_idx:
            st.session_state.active_version_idx = new_idx
            active_content = versions[new_idx]["content"]
            st.session_state.edited_resume = active_content
            st.session_state.enhanced_resume = active_content
            st.session_state._last_enhanced = active_content
            st.session_state.revision_version += 1
            st.session_state.ats_result = None
            st.rerun()
    else:
        if versions:
            st.caption(f"🔵 {versions[0]['label']} ({versions[0]['timestamp']})")

    if st.session_state.get("_last_enhanced") != st.session_state.enhanced_resume:
        st.session_state.edited_resume = st.session_state.enhanced_resume
        st.session_state._last_enhanced = st.session_state.enhanced_resume

    st.caption("Edit directly below before exporting.")
    edited = st.text_area("Edit resume", label_visibility="collapsed",
                          value=st.session_state.edited_resume, height=500,
                          key=f"editable_output_{st.session_state.revision_version}",
                          help="Edit here — changes apply to exports.")
    st.session_state.edited_resume = edited

    col1, col2, col3, _ = st.columns([1, 1, 1, 3])
    with col1:
        st.download_button("⬇ Download .txt", data=edited,
                           file_name=f"enhanced_resume_{date.today()}.txt", mime="text/plain", use_container_width=True)
    with col2:
        st.download_button("⬇ Download .pdf", data=generate_pdf(edited),
                           file_name=f"enhanced_resume_{date.today()}.pdf", mime="application/pdf", use_container_width=True)
    with col3:
        if st.button("↺ Reset edits", use_container_width=True):
            active = versions[st.session_state.active_version_idx]["content"] if versions else st.session_state.enhanced_resume
            st.session_state.edited_resume = active; st.rerun()
else:
    st.markdown('<div class="output-box">Your enhanced resume will appear here after clicking Enhance Resume.</div>', unsafe_allow_html=True)

# ── Skills Gap Analysis ───────────────────────────────────────────────────────
if st.session_state.skills_gap and st.session_state.enhanced_resume:
    st.divider()
    st.subheader("🔍 Skills Gap Analysis")
    st.caption("Based on your enhanced resume vs the JD requirements.")

    raw_gap = st.session_state.skills_gap
    def parse_gap_section(raw, label):
        m = re.search(label + r":\s*(.+?)(?=\n[A-Z]+:|$)", raw, re.S)
        if not m: return []
        return [l.strip().lstrip("-\u2022 ").strip() for l in m.group(1).strip().split("\n")
                if l.strip() and l.strip() not in ("-", "\u2022")]

    have    = parse_gap_section(raw_gap, "HAVE")
    missing = parse_gap_section(raw_gap, "MISSING")
    irrel   = parse_gap_section(raw_gap, "IRRELEVANT")

    col_h, col_m, col_i = st.columns(3)
    with col_h:
        st.markdown("**✅ You Have**")
        for s in have:
            st.markdown(f"- {s}")
        if not have: st.caption("None identified")
    with col_m:
        st.markdown("**❌ Missing / Gaps**")
        for s in missing:
            st.markdown(f"- {s}")
        if not missing: st.caption("None identified")
    with col_i:
        st.markdown("**⚪ Irrelevant to Role**")
        for s in irrel:
            st.markdown(f"- {s}")
        if not irrel: st.caption("None identified")

# ── Follow-up Revision Chat ────────────────────────────────────────────────────
if st.session_state.enhanced_resume:
    st.divider()
    st.subheader("💬 Request Further Revisions")
    st.caption("Ask for specific changes to the enhanced resume. Each revision updates the output above.")

    # Show revision history
    if st.session_state.revision_history:
        with st.expander(f"📜 Revision history ({len(st.session_state.revision_history)} revision(s))", expanded=False):
            for i, entry in enumerate(st.session_state.revision_history, 1):
                st.markdown(f"**Revision {i}:** {entry['request']}")
                st.caption(f"Applied at {entry['timestamp']}")

    revision_input = st.text_area(
        "What would you like to change?",
        placeholder="e.g. 'Make the summary more focused on automation engineering' or 'Strengthen the bullet points in the most recent job' or 'Remove mention of X and replace with Y'",
        height=100,
        key="revision_input",
    )

    col_rev, col_reset_rev, _ = st.columns([1, 1, 4])
    with col_rev:
        apply_revision = st.button("✏️ Apply Revision", use_container_width=True,
                                   disabled=not revision_input.strip())
    with col_reset_rev:
        if st.button("↺ Clear History", use_container_width=True):
            st.session_state.revision_history = []
            st.session_state.edited_resume = st.session_state.enhanced_resume
            st.rerun()

    if apply_revision and revision_input.strip():
        with st.spinner("Applying revision…"):
            try:
                current = st.session_state.edited_resume or st.session_state.enhanced_resume
                rev_content = (
                    f"## Current Resume\n\n{current}"
                    f"\n\n---\n\n## Job Description\n\n{st.session_state.jd_text}"
                    f"\n\n---\n\n## Revision Request\n\n{revision_input}"
                    f"\n\n---\n\n## IMPORTANT INSTRUCTION\n\n"
                    f"Integrate all requested changes directly into the resume body. "
                    f"Do not append any separate section, note, or suggestion at the end. "
                    f"If a skill is requested, add it to the Skills section. "
                    f"Output only the complete revised resume with all changes integrated."
                )
                rev_resp = openai.OpenAI(api_key=OPENAI_API_KEY).chat.completions.create(
                    model=model, temperature=0.5, max_tokens=3500,
                    messages=[
                        {"role": "system", "content": REVISION_PROMPT},
                        {"role": "user", "content": rev_content}
                    ]
                )
                revised = rev_resp.choices[0].message.content.strip()

                # Store revision as a new version
                rev_num = len(st.session_state.resume_versions)
                ts_rev = datetime.now().strftime("%Y-%m-%d %H:%M")
                st.session_state.resume_versions.append({
                    "label": f"Revision {rev_num}: {revision_input[:50]}{'…' if len(revision_input) > 50 else ''}",
                    "content": revised,
                    "timestamp": ts_rev,
                })
                new_idx = len(st.session_state.resume_versions) - 1
                st.session_state.active_version_idx = new_idx
                st.session_state.enhanced_resume = revised
                st.session_state.edited_resume = revised
                st.session_state._last_enhanced = revised
                st.session_state.revision_version += 1
                st.session_state.ats_result = None

                st.session_state.revision_history.append({
                    "request": revision_input,
                    "timestamp": ts_rev,
                })

                # Re-run skills gap on the revised resume so it reflects current state
                try:
                    gap_content = f"## Resume\n\n{revised}\n\n---\n\n## JD\n\n{st.session_state.jd_text}"
                    gr = openai.OpenAI(api_key=OPENAI_API_KEY).chat.completions.create(
                        model="gpt-4o-mini", temperature=0.0, max_tokens=800,
                        messages=[{"role":"system","content":SKILLS_GAP_PROMPT},
                                  {"role":"user","content":gap_content}])
                    st.session_state.skills_gap = gr.choices[0].message.content.strip()
                except Exception:
                    pass  # Keep old skills gap if update fails

                st.success("✅ Revision applied — skills gap updated. Run ATS Score to re-evaluate.")
                st.rerun()
            except Exception as e:
                st.error(f"Revision failed: {e}")

# ── ATS Score display ──────────────────────────────────────────────────────────
# Button-triggered scoring of enhanced/revised resume
if st.session_state.enhanced_resume:
    st.divider()
    col_ats_btn, col_ats_info = st.columns([1, 5])
    with col_ats_btn:
        run_ats = st.button("🎯 Run ATS Score", use_container_width=True,
                            help="Score the current enhanced resume against the JD.")
    with col_ats_info:
        if st.session_state.ats_result:
            st.caption("ATS score shown below. Click again after revisions to re-evaluate.")
        else:
            st.caption("Click to score your enhanced resume against the JD at any time — including after revisions.")

    if run_ats:
        current_resume = st.session_state.edited_resume or st.session_state.enhanced_resume
        # Cache key = hash of resume content + JD — same content always returns same score
        cache_key = hashlib.md5((current_resume + st.session_state.jd_text).encode()).hexdigest()
        if cache_key in st.session_state.ats_score_cache:
            st.session_state.ats_result = st.session_state.ats_score_cache[cache_key]
            st.info("ℹ️ Score retrieved from cache — same resume content returns the same score.")
            st.rerun()
        else:
            with st.spinner("Scoring enhanced resume…"):
                try:
                    ar = openai.OpenAI(api_key=OPENAI_API_KEY).chat.completions.create(
                        model="gpt-4o-mini", temperature=0.0, max_tokens=600,
                        messages=[{"role":"system","content":ENHANCED_ATS_PROMPT},
                                  {"role":"user","content":(
                                      f"## Enhanced Resume\n\n{current_resume}"
                                      f"\n\n---\n\n## JD\n\n{st.session_state.jd_text}"
                                  )}])
                    result = ar.choices[0].message.content.strip()
                    st.session_state.ats_result = result
                    st.session_state.ats_score_cache[cache_key] = result
                    st.rerun()
                except Exception as e:
                    st.error(f"ATS scoring failed: {e}")


def parse_ats(raw: str) -> dict:
    def pf(lbl):
        m = re.search(rf"{lbl}:\s*([\d.]+)", raw)
        return float(m.group(1)) if m else None
    ms = re.search(r"SUMMARY:\s*(.+?)(?=IMPROVEMENTS:|$)", raw, re.S)
    mi = re.search(r"IMPROVEMENTS:\s*(.+?)$", raw, re.S)
    return {"total":pf("TOTAL_SCORE"),"kw":pf("KEYWORD_MATCH"),"exp":pf("RELEVANCE_OF_EXPERIENCE"),
            "qual":pf("QUALIFICATIONS_MATCH"),"clarity":pf("RESUME_CLARITY"),"formatting":pf("ATS_FORMATTING"),
            "summary":ms.group(1).strip() if ms else "","improve":mi.group(1).strip() if mi else ""}

def badge(total, label):
    ok = total >= 60
    c = "#2d5a3d" if ok else "#c0392b"; bg = "#e8f2eb" if ok else "#fde8e8"
    return f'<div style="text-align:center;background:{bg};border:2px solid {c};border-radius:12px;padding:12px 20px"><div style="font-size:11px;color:{c};font-weight:600;margin-bottom:4px">{label}</div><div style="font-size:2.2rem;font-weight:800;color:{c};line-height:1">{int(total)}</div><div style="font-size:11px;color:{c}">/ 100</div><div style="font-size:11px;font-weight:700;color:{c};margin-top:4px">{"✅ PASS" if ok else "❌ BELOW PASSING"}</div></div>'

def bars(scores, base=None):
    cats = [("Keyword Match","kw",30),("Relevance of Experience","exp",25),
            ("Qualifications Match","qual",20),("Resume Clarity","clarity",15),("ATS Formatting","formatting",10)]
    html = ""
    for name, key, mx in cats:
        s = scores.get(key)
        if s is None: continue
        pct = int(s/mx*100)
        bc = "#2d5a3d" if pct>=60 else "#e67e22" if pct>=40 else "#c0392b"
        delta = ""
        if base and (b := base.get(key)) is not None:
            d = s - b; arrow = "▲" if d>0 else ("▼" if d<0 else "–")
            dc = "#2d5a3d" if d>0 else ("#c0392b" if d<0 else "#888")
            delta = f'<span style="color:{dc};font-size:12px;margin-left:8px">{arrow} {abs(d):.0f}</span>'
        html += f'<div style="margin-bottom:12px"><div style="display:flex;justify-content:space-between;font-size:13px;margin-bottom:3px"><span>{name}</span><span style="font-weight:600">{int(s)}/{mx}{delta}</span></div><div style="background:#e0dbd3;border-radius:6px;height:10px"><div style="width:{pct}%;background:{bc};border-radius:6px;height:10px"></div></div></div>'
    return html

if st.session_state.ats_result:
    st.divider(); st.subheader("🎯 ATS Score")
    st.caption("Passing score is 60/100 — Singapore ATS standards and SkillsFuture alignment.")
    es = parse_ats(st.session_state.ats_result)
    bs = parse_ats(st.session_state.baseline_ats_result) if st.session_state.baseline_ats_result else None
    if es["total"] is not None:
        if bs and bs["total"] is not None:
            imp = es["total"] - bs["total"]
            ic = "#2d5a3d" if imp > 0 else "#c0392b"
            sign = "+" if imp > 0 else ""
            c1, c2, c3 = st.columns(3)
            c1.markdown(badge(bs["total"], "BASELINE RESUME"), unsafe_allow_html=True)
            c2.markdown(badge(es["total"], "ENHANCED RESUME"), unsafe_allow_html=True)
            c3.markdown(f'<div style="text-align:center;border:2px solid {ic};border-radius:12px;padding:12px 20px"><div style="font-size:11px;color:{ic};font-weight:600;margin-bottom:4px">IMPROVEMENT</div><div style="font-size:2.2rem;font-weight:800;color:{ic};line-height:1">{sign}{imp:.0f}</div><div style="font-size:11px;color:{ic}">points</div></div>', unsafe_allow_html=True)
        else:
            st.markdown(badge(es["total"], "ENHANCED RESUME"), unsafe_allow_html=True)
        st.write("")
        with st.expander(f"📊 View score breakdown{' *(▲/▼ vs baseline)*' if bs else ''}"):
            st.markdown(bars(es, bs), unsafe_allow_html=True)
            if es["summary"]: st.markdown("**Summary**"); st.info(es["summary"])
            if es["improve"]:
                st.markdown("**How to improve**")
                for l in es["improve"].split("\n"):
                    l = l.strip().lstrip("-•* ").strip()
                    if l: st.markdown(f"- {l}")

st.markdown('<div class="warn-box">⚠️ <strong>Always review the enhanced resume carefully.</strong> AI may occasionally embellish skills or experience — validate before applying.</div>', unsafe_allow_html=True)

if st.session_state.enhanced_resume:
    with st.expander("🔍 Debug — verify ATS inputs & raw scores"):
        st.markdown("**Baseline resume (full text sent to ATS scorer):**")
        st.text_area("Baseline", value=st.session_state.resume_text, height=200,
                     key="debug_baseline", disabled=True, label_visibility="collapsed")
        st.caption(f"{len(st.session_state.resume_text.split())} words")

        st.markdown("**Enhanced resume (full text sent to ATS scorer):**")
        st.text_area("Enhanced", value=st.session_state.enhanced_resume, height=200,
                     key="debug_enhanced", disabled=True, label_visibility="collapsed")
        st.caption(f"{len(st.session_state.enhanced_resume.split())} words")

        st.markdown("**Raw baseline ATS response:**")
        st.code(st.session_state.baseline_ats_result or "None", language="text")
        st.markdown("**Raw enhanced ATS response:**")
        st.code(st.session_state.ats_result or "None", language="text")

        same = st.session_state.resume_text == st.session_state.enhanced_resume
        (st.error if same else st.success)(
            "⚠️ Resumes are IDENTICAL — enhancer may not have run correctly." if same
            else f"✅ Resumes differ — enhancement applied ({len(st.session_state.resume_text.split())} → {len(st.session_state.enhanced_resume.split())} words)."
        )

# ── Application Tracker ────────────────────────────────────────────────────────
st.divider(); st.subheader("📋 Application Tracker")
st.caption("Entries are staged in session. Use the save form to push confirmed entries to Supabase.")

def get_defaults() -> dict:
    df = st.session_state.tracker_edit
    if df.empty:
        return {
            "company": parse_jd_field(st.session_state.jd_text, [r"(?i)company name", r"(?i)company"]),
            "role":    parse_jd_field(st.session_state.jd_text, [r"(?i)job title", r"(?i)role", r"(?i)position"]),
            "status":  "Not Applied"
        }
    last = df.iloc[-1]
    return {"company": str(last.get("Company","NA") or "NA"),
            "role":    str(last.get("Role","NA") or "NA"),
            "status":  str(last.get("Status","Not Applied") or "Not Applied")}

with st.expander("💾 Select & save entry to Supabase"):
    df_s = st.session_state.tracker_edit
    if df_s.empty:
        st.info("No entries yet — enhance a resume first.")
    else:
        labels = [f"#{i+1} — {r.get('Company','NA')} | {r.get('Role','NA')} | {r.get('Date Applied','')}"
                  for i, r in df_s.iterrows()]
        sel = st.selectbox("Select entry:", labels, index=len(labels)-1, key="save_row_select")
        si  = labels.index(sel)
        row = df_s.iloc[si]

        # Version key based on selected index — forces Streamlit to re-render
        # inputs with the correct row data when selection changes
        vk = f"_v{si}"

        if st.button("🗑️ Delete this entry"):
            st.session_state.tracker_edit = df_s.drop(index=df_s.index[si]).reset_index(drop=True); st.rerun()

        st.divider(); st.markdown("**Review & edit before saving:**")
        c1, c2 = st.columns(2)
        with c1:
            nc = st.text_input("Company", value=str(row.get("Company","NA") or "NA"), key=f"new_company{vk}")
            nr = st.text_input("Role",    value=str(row.get("Role","NA") or "NA"),    key=f"new_role{vk}")
        with c2:
            try:   dv = pd.to_datetime(row.get("Date Applied", str(date.today()))).date()
            except: dv = date.today()
            nd = st.date_input("Date Applied", value=dv, key=f"new_date{vk}")
            sv = str(row.get("Status","Not Applied") or "Not Applied")
            ns = st.selectbox("Status", STATUS_OPTIONS,
                              index=STATUS_OPTIONS.index(sv) if sv in STATUS_OPTIONS else 0, key=f"new_status{vk}")

        st.info(f"📋 **Will save:** {nc} | {nr} | {nd} | {ns}")
        if st.button("✅ Confirm & Save to Supabase"):
            rid, err = sb_insert(nc, nr, nd, ns)
            if err:
                st.error(f"❌ {err}")
            else:
                idx = df_s.index[si]
                for col, val in [("Company",nc),("Role",nr),("Date Applied",str(nd)),("Status",ns)]:
                    st.session_state.tracker_edit.at[idx, col] = val
                if rid: st.session_state.tracker_edit.at[idx, "id"] = rid
                st.success(f"✅ Saved: {nc} — {nr}"); st.rerun()

display_df = st.session_state.tracker_edit.drop(columns=["id"], errors="ignore").copy()
if "Date Applied" in display_df.columns:
    display_df["Date Applied"] = pd.to_datetime(display_df["Date Applied"], errors="coerce").dt.date
display_df.insert(0, "#", range(1, len(display_df)+1))

edited_t = st.data_editor(display_df, num_rows="dynamic", use_container_width=True, disabled=["#"],
    column_config={
        "#": st.column_config.NumberColumn("#", width="small"),
        "Company": st.column_config.TextColumn("Company"),
        "Role": st.column_config.TextColumn("Role"),
        "Date Applied": st.column_config.DateColumn("Date Applied", format="YYYY-MM-DD"),
        "Status": st.column_config.SelectboxColumn("Status", options=STATUS_OPTIONS, required=True),
    }, key="tracker_editor")

edn = edited_t.drop(columns=["#"], errors="ignore")
ddn = display_df.drop(columns=["#"], errors="ignore")
if not edn.equals(ddn):
    idc = st.session_state.tracker_edit["id"] if "id" in st.session_state.tracker_edit.columns else pd.Series(dtype=str)
    for i, row in edn.iterrows():
        sb_update(str(idc.iloc[i] if i < len(idc) else ""),
                  str(row.get("Company","")), str(row.get("Role","")),
                  str(row.get("Date Applied","")), str(row.get("Status","")))
    eid = edn.copy()
    if "id" in st.session_state.tracker_edit.columns:
        eid["id"] = st.session_state.tracker_edit["id"].reindex(edn.index).values
    st.session_state.tracker_edit = eid

with st.columns([1, 5])[0]:
    st.download_button("⬇ Export Excel", data=export_excel(edited_t),
                       file_name=f"applications_{date.today()}.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       use_container_width=True)

st.caption("✅ Connected to Supabase" if SUPABASE_URL and SUPABASE_ANON_KEY else "⚠️ Supabase not configured — entries stored in session only.")